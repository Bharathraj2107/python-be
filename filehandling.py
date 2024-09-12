# 2 types of files
#system files
#stream files
#file=open('C:/users/admin/desktop/greet.txt','w')#we can specify any path were we want to create the file
#greet.txt is the file u want to create and w is used for write mode
# file=open("greet.txt",'w')#here even if the file not present the file will be created
# file.write("Hello,welcome to the world of file handling \n")
# file.write("This is the write command\n")
# file.write("It allows to write in a particular lines files \n")
# file.close()

#read the file

# file='greet.txt'
# filehandle=open(file,'r')
# filedata=filehandle.read()
# print(filedata)

# fruits=["apple \n","orange \n","grapes \n"]
# my_file=open('fruits.txt',"w")
# my_file.writelines(fruits)
# my_file.close()

# my_file=open("fruits.txt","a+")
# my_file.write("\nstraberry")
# filehandle=open("fruits.txt",'r')
# my_file=filehandle.read()
# print(my_file)

# fruits=["\nBanana","\nAvocado","\nFigs"]
# my_file=open("fruits.txt","a+")
# my_file.writelines(fruits)
# filehandle=open("fruits.txt",'r')
# my_file=filehandle.read()
# print(my_file)

#open file
# f=open("fruits.txt","r+")
# #absolute file positioning
# f.seek(0)#to bring it to the start of the line
# #to erase all the data
# f.truncate()
# #
# f=open("fruits.txt","r")
# my_file=f.read()
# print(my_file)
# f.close()

# import os
# os.rename("c:/some path/greet.txt,( wher the file is located)","same path last change the name of file/my_test.txt")#here greet .txt is changed to my_test.txt
# os.remove("fruits.txt")#to remove the file
# #binadry file encode and decode
# my_file=open("specify path","wb+")
# message="hello python"
# file_encode=message.encode("ASCII")
# my_file.write(file_encode)
# my_file.seek(0)
# bdata=my_file.read()
# print("binary data",bdata)
# ntext=bdata.decode("ASCII")
# print(ntext)
# print("normal data:",ntext)

#xl file
import openpyxl
# path="E:\\pythonProject5\\Book1.xlsx"#This is the path were file is stored
# workbook=openpyxl.load_workbook(path)#this is used to load the excel file
# sheet=workbook.active#to make the Excel sheet read from starting we use this
# for r in range(1,6):
#     for c in range(1,6):
#         sheet.cell(row=r,column=c).value=" Hi, welcome"#here the intersection of row and column forms a cell to insert we use .value
# workbook.save(path)
#to read we use this
path="E:\\pythonProject5\\Book1.xlsx"#This is the path were file is stored
workbook=openpyxl.load_workbook(path)#this is used to load the excel file
sheet=workbook.active#workbook.get_sheet_by_name("sheet1")#This which page we need that sheet we specify
rows=sheet.max_row#it gives number of rows
cols=sheet.max_column#it gives number of columns
print(rows)
print(cols)
for r in range(1,rows+1):
    for c in range(1,cols+1):
        print(sheet.cell(row=r,column=c).value,end=" ")
    print()